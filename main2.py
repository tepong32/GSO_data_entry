import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from ttkthemes import ThemedTk
import openpyxl
import pandas as pd



file_path = ""  # Global variable to store the file path
current_sheet = ""  # Global variable to store the selected sheet name
widget_name_mapping = {}  # Global variable to store the mapping of widgets
column_to_widget_mapping = {} 

### this first asks the user which excel file to load
def load_excel_file():
    global file_path, current_sheet
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xlsm")])
    if file_path:
        current_sheet = ""  # Clear the current sheet when loading a new file
        load_data(file_path)

### you can use this instead of the above code chunk to directly loads the db.xlsx file upon app startup
### comment-out the previous and un-comment this chunk
# def load_excel_file():
#     global current_sheet, file_path
#     file_path = "db.xlsx"  # Directly set the file path to "db.xlsx" in the root folder
#     current_sheet = ""  # Clear the current sheet when loading a new file
#     load_data(file_path)

def load_data(file_path):
    global current_sheet, widget_name_mapping
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet_names = workbook.sheetnames
        
        if not current_sheet or current_sheet not in sheet_names:
            # If current_sheet is empty or not in sheet_names, set it to the first sheet
            current_sheet = sheet_names[0]
            
        # Update the Combobox with sheet names
        sheet_dropdown["values"] = sheet_names
        sheet_dropdown.set(current_sheet)  # Set the selected sheet in the Combobox

        sheet = workbook[current_sheet]
        list_values = list(sheet.iter_rows(values_only=True))
        columns = list_values[0]  # Extract the column headings from the first row of the selected sheet

        # Update the TreeView's headings with the column headings
        update_treeview_headings(columns)

        widget_name_mapping = {}
        column_to_widget_mapping = {}  # Reverse mapping: Column headings to widget names

        for idx, col_name in enumerate(columns):
            if idx < 6:
                widget_name = "widget_" + str(idx)
                widget_name_mapping[widget_name] = col_name
                column_to_widget_mapping[col_name] = widget_name.lower()  # Create the reverse mapping with lowercase
                # Update the label for the corresponding widget
                widgets_frame.grid_slaves(row=0, column=idx)[0].config(text=col_name)

        # Insert the data into the TreeView
        for value_tuple in list_values[1:]:
            treeView.insert('', tk.END, values=value_tuple)

        return sheet_names, column_to_widget_mapping  # Return both sheet names list and the reverse mapping
    
    except openpyxl.utils.exceptions.InvalidFileException:
        print("Invalid Excel file. Please select a valid Excel file.")

        
def on_sheet_select(event):
    global current_sheet
    selected_sheet = sheet_dropdown.get()
    if selected_sheet != current_sheet:
        current_sheet = selected_sheet
        load_data(file_path)  # Update the TreeView when the selected sheet changes

def update_treeview_headings(columns):
    # Clear previous TreeView columns and headings
    treeView.delete(*treeView.get_children())
    treeView["columns"] = columns

    for col_name in columns:
        treeView.heading(col_name, text=col_name)
        treeView.column(col_name, width=100)  # Set a default width for each column (you can adjust it as needed)

def insert_row():
    if not column_to_widget_mapping:
        print("Error: No mapping available.")
    # Retrieve data from the widgets
    r1 = month.get()
    r2 = obr_number.get()
    r3 = category_combobox.get()
    r4 = sub_category_combobox.get()
    r5 = brand.get()
    r6 = price.get()
    r7 = notes.get()

    # Get the selected sheet from the workbook
    path = file_path
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[current_sheet]
    
    # Insert the row to the selected sheet in the workbook
    row_values = [r1, r2, r3, r4, r5, r6, r7]
    sheet.append(row_values)
    workbook.save(path)

    # displaying the inserted row on the UI (treeView)
    treeView.insert('', tk.END, values=row_values)
    
    # Get the ID of the last inserted row in the TreeView
    last_row_id = treeView.get_children()[-1]

    # clear the values after inserting the new row then resetting the values to default
    month.delete(0, "end")
    obr_number.delete(0, "end")
    category_combobox.set("Select Category")
    sub_category_combobox.delete(0, "end")
    brand.delete(0, "end")
    price.delete(0, "end")
    notes.delete(0, "end")

    # Highlight or select the last inserted row in the TreeView
    treeView.selection_set(last_row_id)
    # Scroll the TreeView to make sure the last inserted row is visible
    treeView.see(last_row_id)
    
    # returns the focus to the "month" widget after inserting the new row
    month.focus_set()


# Create the main window
root = ThemedTk(theme="radiance") #tk.Tk()
root.title('GSO Records App by tEppy™')
### this two lines below set the app to fullscreen
# root.overrideredirect(True)
# root.geometry(f"{root.winfo_screenwidth()}x{root.winfo_screenheight()}+0+0")

# Update labels of widgets based on column headings and get the widget_name_mapping
sheet_names = load_data(file_path)  # Get the sheet names list

################## Main Frame ##################
'''
    outer_frame.pack() makes the app responsive.
    Since this is the main widget, adjusting the size of the UI/app will keep
    the components centered
'''
outer_frame = ttk.Frame(root) # parent widget
outer_frame.pack()

# Select File Button
btn_load = ttk.Button(outer_frame, text="Load Records", command=load_excel_file, takefocus=0)
btn_load.grid(row=0, column=0, padx=5, pady=(0, 5), sticky="e")

################## Widgets Frame ##################
''' 
    The user-input form area.
    All other ttk widgets should have "widgets_frame" as their parent
    since they are supposed to be "grouped-together"... except from, of course, the
    widgets_frame for its root will be the outer_frame (see above).
'''
### col0, row0 of the root_frame : Enclosure Widget for all other input widgets
widgets_frame = ttk.LabelFrame(outer_frame, text="Widgets Frame")
widgets_frame.grid(row=1, column=0, padx=20, pady=10) # padding on x & y axis

w00 = ttk.Label(widgets_frame, )
w00.grid(row=0, column=0)

w01 = ttk.Label(widgets_frame, )
w01.grid(row=0, column=1)

w02 = ttk.Label(widgets_frame, )
w02.grid(row=0, column=2)

w03 = ttk.Label(widgets_frame, )
w03.grid(row=0, column=3)

w04 = ttk.Label(widgets_frame, )
w04.grid(row=0, column=4)

w05 = ttk.Label(widgets_frame, )
w05.grid(row=0, column=5)

w06 = ttk.Label(widgets_frame, )
w06.grid(row=0, column=6)

w07 = ttk.Label(widgets_frame, text="Qty")
w07.grid(row=3, column=1)
w08 = ttk.Label(widgets_frame, text="Office")
w08.grid(row=3, column=2)
w09 = ttk.Label(widgets_frame, text="w9")
w09.grid(row=3, column=3)
w10 = ttk.Label(widgets_frame, text="w10")
w10.grid(row=3, column=4)
w11 = ttk.Label(widgets_frame, text="w11")
w11.grid(row=3, column=5)

# Read Excel data from a specific sheet (replace 'Sheet1' with your actual sheet name)
df = pd.read_excel('db.xlsx', sheet_name='Lists(DoNotSelectThis)')
# setting variables for dropdown options
asset_type_list = df['AssetType'].tolist()
chairs_list = df['List_Chairs'].tolist()
departments_list = df['Departments'].tolist()
kitchen_list = df['List_Kitchen'].tolist()
others_list = df['List_Others'].tolist()
peripherals_list = df['List_ComputerPeripherals'].tolist()
racks_list = df['List_Racks'].tolist()
tables_list = df['List_Tables'].tolist()


### actual widgets
month = ttk.Entry(widgets_frame, width=7)
# check_date.insert(0, "help text here")
# check_date.bind("<FocusIn>", lambda e: check_date.delete('0', 'end'))
month.grid(row=2, column=0, padx=5, pady=(0,5), sticky="ew")

obr_number = ttk.Entry(widgets_frame,width=13)
obr_number.grid(row=2, column=1, padx=5, pady=(0,5), sticky="ew")

# category_list = asset_type_list
# category_dropdown = ttk.Combobox(widgets_frame, values=category_list, state='readonly')
# category_dropdown.current(0) # default selected value
# category_dropdown.grid(row=2, column=2, padx=5, pady=(0,5), sticky="ew")

def on_category_selected(event):
    selected_category = category_var.get()
    if selected_category == "Computer":
        sub_category_combobox['values'] = peripherals_list
        sub_category_combobox.set(peripherals_list[0])  # Set default value
        sub_category_combobox['state'] = 'readonly'
        sub_category_combobox.grid(row=2, column=3, padx=5, pady=(0, 5), sticky="ew")
    elif selected_category == "Kitchen-related":
        sub_category_combobox['values'] = kitchen_list
        sub_category_combobox.set(kitchen_list[0])  # Set default value
        sub_category_combobox['state'] = 'readonly'
        sub_category_combobox.grid(row=2, column=3, padx=5, pady=(0, 5), sticky="ew")
    elif selected_category == "Table":
        sub_category_combobox['values'] = tables_list
        sub_category_combobox.set(tables_list[0])  # Set default value
        sub_category_combobox['state'] = 'readonly'
        sub_category_combobox.grid(row=2, column=3, padx=5, pady=(0, 5), sticky="ew")
    elif selected_category == "Chair":
        sub_category_combobox['values'] = chairs_list
        sub_category_combobox.set(chairs_list[0])  # Set default value
        sub_category_combobox['state'] = 'readonly'
        sub_category_combobox.grid(row=2, column=3, padx=5, pady=(0, 5), sticky="ew")
    elif selected_category == "Rack":
        sub_category_combobox['values'] = racks_list
        sub_category_combobox.set(racks_list[0])  # Set default value
        sub_category_combobox['state'] = 'readonly'
        sub_category_combobox.grid(row=2, column=3, padx=5, pady=(0, 5), sticky="ew")
    elif selected_category == "Others":
        sub_category_combobox['values'] = others_list
        sub_category_combobox.set(others_list[0])  # Set default value
        sub_category_combobox['state'] = 'readonly'
        sub_category_combobox.grid(row=2, column=3, padx=5, pady=(0, 5), sticky="ew")
    else:
        sub_category_combobox['state'] = 'disabled'
        sub_category_combobox.grid(row=2, column=3, padx=5, pady=(0, 5), sticky="ew")

# Category Combobox
category_var = tk.StringVar()
category_combobox = ttk.Combobox(widgets_frame, textvariable=category_var, values=asset_type_list, state='readonly')
category_combobox.set("Select Category")
category_combobox.grid(row=2, column=2, padx=5, pady=(0, 5), sticky="ew")
category_combobox.bind("<<ComboboxSelected>>", on_category_selected)

# Subcategory Combobox (initially hidden)
sub_category_combobox = ttk.Combobox(widgets_frame, state='readonly')
sub_category_combobox.grid(row=2, column=3, padx=5, pady=(0, 5), sticky="ew")





brand = ttk.Entry(widgets_frame, width=20)
brand.grid(row=2, column=4,padx=5, pady=(0,5), sticky="ew")

price = ttk.Entry(widgets_frame, width=15)
price.grid(row=2, column=5,padx=5, pady=(0,5), sticky="ew")

notes = ttk.Entry(widgets_frame, width=25)
notes.grid(row=2, column=6, padx=5, pady=(0,5), sticky="ew")



qty = ttk.Entry(widgets_frame, width=25)
qty.grid(row=4, column=1, padx=5, pady=(0,5), sticky="ew")
office = ttk.Entry(widgets_frame, width=25)
office.grid(row=4, column=2, padx=5, pady=(0,5), sticky="ew")
w9_entry = ttk.Entry(widgets_frame, width=25)
w9_entry.grid(row=4, column=3, padx=5, pady=(0,5), sticky="ew")


# Select Sheet Button (Inside Widgets Frame)
def select_sheet():
    selected_item = treeView.focus()
    if selected_item:
        item_values = treeView.item(selected_item, "values")
        selected_sheet_name = item_values[0]
        print(f"Selected sheet: {selected_sheet_name}")

# Dropdown menu for selecting sheets
sheet_dropdown = ttk.Combobox(widgets_frame, state="readonly")
sheet_dropdown.grid(row=4, column=5, padx=10, pady=5)
sheet_dropdown.bind("<<ComboboxSelected>>", on_sheet_select)

# Insert Row button (Inside Widgets Frame)
btn_row = ttk.Button(widgets_frame, text="Insert", command=lambda: insert_row(),
                    takefocus=1)
btn_row.grid(row=4, column=3, sticky="nsew")


################## TreeView / Excel LabelFrame ####################################
### This is where the preview of the excel file's data will be displayed

### Display Frame
treeFrame = ttk.Frame(outer_frame, takefocus=0)
treeFrame.grid(row=5, column=0, pady=10)
treeScroll = ttk.Scrollbar(treeFrame)
treeScroll.pack(side="right", fill="y") # this sets the scrollbar to the right side of the frame,
                                        # covering its whole height
cols = ("Month", "OBR #", "Category","Sub-category", "Brand", "Price", "Notes") # column of the preview related to the excel file
treeView = ttk.Treeview(treeFrame, show="headings", 
                        yscrollcommand=treeScroll.set, columns=cols, height=15)
# these set the width of the columns specifically
# COLUMN NAMES SHOULD MATCH THOSE INDICATED ON THE EXCEL FILE
treeView.column("Month", width=100)
treeView.column("OBR #", width=100)
treeView.column("Category", width=100)
treeView.column("Sub-category", width=200)
treeView.column("Brand", width=100)
treeView.column("Price", width=100)
treeView.column("Notes", width=80)
treeView.pack()
treeScroll.config(command=treeView.yview) # this line attaches the treeScroll widget to the treeView, scrolling vertically


# Event Listener function highlighting selected items on the treeView list
def selected():
    print(listbox.get(listbox.curselection()[0]))
    
treeView.bind("<<ListboxSelect>>", lambda x: selected())

# ### switch (dark/light)
# def toggle_mode():
#     if mode_switch.instate(["selected"]):
#         style.theme_use("forest-light")
#     else:
#         style.theme_use("forest-dark")

# mode_switch = ttk.Checkbutton(outer_frame, text="Mode", style="Switch",
#     command=toggle_mode, takefocus=0) # this triggers the toggle_mode function above
# mode_switch.grid(row=6, column=0, padx=5, pady=10, sticky="nsew")

month.focus_set()

root.mainloop()
"""_summary_
"""