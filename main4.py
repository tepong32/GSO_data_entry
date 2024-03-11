### from tutorial to actual app build

from tkinter import *
from tkinter import filedialog
from tkinter import ttk
from ttkthemes import ThemedTk
import openpyxl
import pandas as pd


root = ThemedTk(theme="radiance")
root.title("App by tEppy™")
# root.geometry("1200x800")  # Initial window size
# # root.resizable(width=False, height=False) # fixed-size app window
root.resizable(0, 0)

##### defining functions #####

file_path = ""  # Global variable to store the file path
current_sheet = ""  # Global variable to store the selected sheet name
widget_name_mapping = {}  # Global variable to store the mapping of widgets
column_to_widget_mapping = {} 

### this first asks the user which excel file to load
def load_excel_file():
    global file_path, current_sheet
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xlsm")])
    if file_path:
        current_sheet = ""  # Clear the current sheet when loading a new file
        load_data(file_path)

### you can use this instead of the above code chunk to directly loads the db.xlsx file upon app startup
### comment-out the previous and un-comment this chunk
# def load_excel_file():
#     global current_sheet, file_path
#     file_path = "db.xlsx"  # Directly set the file path to "db.xlsx" in the root folder
#     current_sheet = ""  # Clear the current sheet when loading a new file
#     load_data(file_path)

def load_data(file_path):
    global current_sheet, widget_name_mapping
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet_names = workbook.sheetnames
        
        if not current_sheet or current_sheet not in sheet_names:
            # If current_sheet is empty or not in sheet_names, set it to the first sheet
            current_sheet = sheet_names[0]
            
        # Update the Combobox with sheet names
        sheet_dropdown["values"] = sheet_names
        sheet_dropdown.set(current_sheet)  # Set the selected sheet in the Combobox

        sheet = workbook[current_sheet]
        list_values = list(sheet.iter_rows(values_only=True))
        columns = list_values[0]  # Extract the column headings from the first row of the selected sheet

        # Update the TreeView's headings with the column headings
        update_treeview_headings(columns)

        widget_name_mapping = {}
        column_to_widget_mapping = {}  # Reverse mapping: Column headings to widget names

        for idx, col_name in enumerate(columns):
            if idx < 6:
                widget_name = "widget_" + str(idx)
                widget_name_mapping[widget_name] = col_name
                column_to_widget_mapping[col_name] = widget_name.lower()  # Create the reverse mapping with lowercase
                # Update the label for the corresponding widget
                widgets_frame.grid_slaves(column=0, row=idx)[0].config(text=col_name)

        # Insert the data into the TreeView
        for value_tuple in list_values[1:]:
            treeView.insert('', END, values=value_tuple)

        return sheet_names, column_to_widget_mapping  # Return both sheet names list and the reverse mapping
    
    except openpyxl.utils.exceptions.InvalidFileException:
        print("Invalid Excel file. Please select a valid Excel file.")

        
def on_sheet_select(event):
    global current_sheet
    selected_sheet = sheet_dropdown.get()
    if selected_sheet != current_sheet:
        current_sheet = selected_sheet
        load_data(file_path)  # Update the TreeView when the selected sheet changes

def update_treeview_headings(columns):
    # Clear previous TreeView columns and headings
    treeView.delete(*treeView.get_children())
    treeView["columns"] = columns

    for col_name in columns:
        treeView.heading(col_name, text=col_name)
        treeView.column(col_name, width=100)  # Set a default width for each column (you can adjust it as needed)

def insert_row():
    if not column_to_widget_mapping:
        print("Error: No mapping available.")
    # Retrieve data from the widgets
    r1 = month.get()
    r2 = obr_number.get()
    r3 = category_combobox.get()
    r4 = sub_category_combobox.get()
    r5 = brand.get()
    r6 = price.get()
    r7 = notes.get()

    # Get the selected sheet from the workbook
    path = file_path
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[current_sheet]
    
    # Insert the row to the selected sheet in the workbook
    row_values = [r1, r2, r3, r4, r5, r6, r7]
    sheet.append(row_values)
    workbook.save(path)

    # displaying the inserted row on the UI (treeView)
    treeView.insert('', END, values=row_values)
    
    # Get the ID of the last inserted row in the TreeView
    last_row_id = treeView.get_children()[-1]

    # clear the values after inserting the new row then resetting the values to default
    month.delete(0, "end")
    obr_number.delete(0, "end")
    category_combobox.set("Select Category")
    sub_category_combobox.delete(0, "end")
    brand.delete(0, "end")
    price.delete(0, "end")
    notes.delete(0, "end")

    # Highlight or select the last inserted row in the TreeView
    treeView.selection_set(last_row_id)
    # Scroll the TreeView to make sure the last inserted row is visible
    treeView.see(last_row_id)
    
    # returns the focus to the "month" widget after inserting the new row
    month.focus_set()

outer_frame = ttk.Frame(root)
widgets_frame = ttk.Frame(outer_frame, borderwidth=5, relief="raised") # change height later.


################## Widgets Frame ##################
''' 
    The user-input form area.
    All other ttk widgets should have "widgets_frame" as their parent
    since they are supposed to be "grouped-together"... except from, of course, the
    widgets_frame for its root will be the outer_frame (see above).
'''
### col0, row0 of the root_frame : Enclosure Widget for all other input widgets
widgets_frame = ttk.LabelFrame(outer_frame, text="Widgets Frame")
widgets_frame.grid(row=1, column=0, padx=20, pady=10) # padding on x & y axis

w00 = ttk.Label(widgets_frame, )
w00.grid(column=0, row=0, )

w01 = ttk.Label(widgets_frame, )
w01.grid(column=0, row=1, )

w02 = ttk.Label(widgets_frame, )
w02.grid(column=0, row=2, )

w03 = ttk.Label(widgets_frame, )
w03.grid(column=0, row=3, )

w04 = ttk.Label(widgets_frame, )
w04.grid(column=0, row=4, )

w05 = ttk.Label(widgets_frame, )
w05.grid(column=0, row=5, )

w06 = ttk.Label(widgets_frame, )
w06.grid(column=0, row=6, )

w07 = ttk.Label(widgets_frame, text="Qty")
w07.grid(column=0, row=7, )
w08 = ttk.Label(widgets_frame, text="Office")
w08.grid(column=0, row=8, )
w09 = ttk.Label(widgets_frame, text="w9")
w09.grid(column=0, row=9, )
w10 = ttk.Label(widgets_frame, text="w10")
w10.grid(column=0, row=10, )
w11 = ttk.Label(widgets_frame, text="w11")
w11.grid(column=0, row=11, )

# Read Excel data from a specific sheet (replace 'Sheet1' with your actual sheet name)
df = pd.read_excel('db.xlsx', sheet_name='Lists(DoNotSelectThis)')
# setting variables for dropdown options
asset_type_list = df['AssetType'].tolist()
chairs_list = df['List_Chairs'].tolist()
departments_list = df['Departments'].tolist()
kitchen_list = df['List_Kitchen'].tolist()
others_list = df['List_Others'].tolist()
peripherals_list = df['List_ComputerPeripherals'].tolist()
racks_list = df['List_Racks'].tolist()
tables_list = df['List_Tables'].tolist()


### actual widgets
month = ttk.Entry(widgets_frame, width=7)
# check_date.insert(0, "help text here")
# check_date.bind("<FocusIn>", lambda e: check_date.delete('0', 'end'))
month.grid(column=1, row=0, padx=5, pady=(0,5), sticky="ew")

obr_number = ttk.Entry(widgets_frame,width=13)
obr_number.grid(column=1, row=1, padx=5, pady=(0,5), sticky="ew")

# category_list = asset_type_list
# category_dropdown = ttk.Combobox(widgets_frame, values=category_list, state='readonly')
# category_dropdown.current(0) # default selected value
# category_dropdown.grid(row=2, column=2, padx=5, pady=(0,5), sticky="ew")

def on_category_selected(event):
    selected_category = category_var.get()
    if selected_category == "Computer":
        sub_category_combobox['values'] = peripherals_list
        sub_category_combobox.set(peripherals_list[0])  # Set default value
        sub_category_combobox['state'] = 'readonly'
        sub_category_combobox.grid(row=2, column=3, padx=5, pady=(0, 5), sticky="ew")
    elif selected_category == "Kitchen-related":
        sub_category_combobox['values'] = kitchen_list
        sub_category_combobox.set(kitchen_list[0])  # Set default value
        sub_category_combobox['state'] = 'readonly'
        sub_category_combobox.grid(row=2, column=3, padx=5, pady=(0, 5), sticky="ew")
    elif selected_category == "Table":
        sub_category_combobox['values'] = tables_list
        sub_category_combobox.set(tables_list[0])  # Set default value
        sub_category_combobox['state'] = 'readonly'
        sub_category_combobox.grid(row=2, column=3, padx=5, pady=(0, 5), sticky="ew")
    elif selected_category == "Chair":
        sub_category_combobox['values'] = chairs_list
        sub_category_combobox.set(chairs_list[0])  # Set default value
        sub_category_combobox['state'] = 'readonly'
        sub_category_combobox.grid(row=2, column=3, padx=5, pady=(0, 5), sticky="ew")
    elif selected_category == "Rack":
        sub_category_combobox['values'] = racks_list
        sub_category_combobox.set(racks_list[0])  # Set default value
        sub_category_combobox['state'] = 'readonly'
        sub_category_combobox.grid(row=2, column=3, padx=5, pady=(0, 5), sticky="ew")
    elif selected_category == "Others":
        sub_category_combobox['values'] = others_list
        sub_category_combobox.set(others_list[0])  # Set default value
        sub_category_combobox['state'] = 'readonly'
        sub_category_combobox.grid(row=2, column=3, padx=5, pady=(0, 5), sticky="ew")
    else:
        sub_category_combobox['state'] = 'disabled'
        sub_category_combobox.grid(row=2, column=3, padx=5, pady=(0, 5), sticky="ew")

# Category Combobox
category_var = StringVar()
category_combobox = ttk.Combobox(widgets_frame, textvariable=category_var, values=asset_type_list, state='readonly')
category_combobox.set("Select Category")
category_combobox.grid(column=1, row=2, padx=5, pady=(0,5), sticky="ew")
category_combobox.bind("<<ComboboxSelected>>", on_category_selected)

# Subcategory Combobox (initially hidden)
sub_category_combobox = ttk.Combobox(widgets_frame, state='readonly')
sub_category_combobox.grid(column=1, row=3, padx=5, pady=(0,5), sticky="ew")

brand = ttk.Entry(widgets_frame, width=20)
brand.grid(column=1, row=4, padx=5, pady=(0,5), sticky="ew")

price = ttk.Entry(widgets_frame, width=15)
price.grid(column=1, row=5, padx=5, pady=(0,5), sticky="ew")

notes = ttk.Entry(widgets_frame, width=25)
notes.grid(column=1, row=6, padx=5, pady=(0,5), sticky="ew")

qty = ttk.Entry(widgets_frame, width=25)
qty.grid(column=1, row=7, padx=5, pady=(0,5), sticky="ew")
office = ttk.Entry(widgets_frame, width=25)
office.grid(column=1, row=8, padx=5, pady=(0,5), sticky="ew")
w9_entry = ttk.Entry(widgets_frame, width=25)
w9_entry.grid(column=1, row=9, padx=5, pady=(0,5), sticky="ew")


# Select Sheet Button (Inside Widgets Frame)
def select_sheet():
    selected_item = treeView.focus()
    if selected_item:
        item_values = treeView.item(selected_item, "values")
        selected_sheet_name = item_values[0]
        print(f"Selected sheet: {selected_sheet_name}")

# Dropdown menu for selecting sheets
sheet_dropdown = ttk.Combobox(widgets_frame, state="readonly")
sheet_dropdown.grid(column=1, row=10, padx=5, pady=(0,5), sticky="ew")
sheet_dropdown.bind("<<ComboboxSelected>>", on_sheet_select)

# Insert Row button (Inside Widgets Frame)
btn_row = ttk.Button(widgets_frame, text="Insert", command=lambda: insert_row(),
                    takefocus=1)
btn_row.grid(column=2, row=11, padx=5, pady=(0,5), sticky="ew")


################## TreeView / Excel LabelFrame ####################################
### This is where the preview of the excel file's data will be displayed

### Display Frame
treeFrame = ttk.Frame(outer_frame, borderwidth=5, relief="flat", takefocus=0)
treeScroll = ttk.Scrollbar(treeFrame)
treeScroll.pack(side="right", fill="y") # this sets the scrollbar to the right side of the frame,
                                        # covering its whole height
cols = ("Month", "OBR #", "Category","Sub-category", "Brand", "Price", "Notes") # column of the preview related to the excel file
treeView = ttk.Treeview(treeFrame, show="headings", 
                        yscrollcommand=treeScroll.set, columns=cols, height=15)
# these set the width of the columns specifically
# COLUMN NAMES SHOULD MATCH THOSE INDICATED ON THE EXCEL FILE
treeView.column("Month", width=60)
treeView.column("OBR #", width=60)
treeView.column("Category", width=60)
treeView.column("Sub-category", width=60)
treeView.column("Brand", width=60)
treeView.column("Price", width=60)
treeView.column("Notes", width=100)
treeView.pack()
treeScroll.config(command=treeView.yview) # this line attaches the treeScroll widget to the treeView, scrolling vertically


# Event Listener function highlighting selected items on the treeView list
def selected():
    print(listbox.get(listbox.curselection()[0]))
    
treeView.bind("<<ListboxSelect>>", lambda x: selected())

# ### switch (dark/light)
# def toggle_mode():
#     if mode_switch.instate(["selected"]):
#         style.theme_use("forest-light")
#     else:
#         style.theme_use("forest-dark")

# mode_switch = ttk.Checkbutton(outer_frame, text="Mode", style="Switch",
#     command=toggle_mode, takefocus=0) # this triggers the toggle_mode function above
# mode_switch.grid(row=6, column=0, padx=5, pady=10, sticky="nsew")

month.focus_set()



# Widget placements (grid for the root frame)
outer_frame.grid(column=0, row=0) # main
widgets_frame.grid(column=0, row=0) # 1st frame spanning 3colsx10rows
treeFrame.grid(column=1, row=0, pady=10)

# Select File Button
btn_load = ttk.Button(outer_frame, text="Load Records", command=load_excel_file, takefocus=0)
btn_load.grid(column=1, row=2, padx=5, pady=(0, 5), sticky="w")

exit_button = ttk.Button(outer_frame, text="Exit", command=quit, takefocus=0)
exit_button.grid(column=1, row=3, padx=5, pady=(0, 5), sticky="w")



################################################################## these lines can be used to set vars to "still useful?" and "returned?". Do not remove these.
onevar = BooleanVar(value=True) 
twovar = BooleanVar(value=False)
threevar = BooleanVar(value=True)

one = ttk.Checkbutton(widgets_frame, text="One", variable=onevar, onvalue=True)
two = ttk.Checkbutton(widgets_frame, text="Two", variable=twovar, onvalue=True)
three = ttk.Checkbutton(widgets_frame, text="Three", variable=threevar, onvalue=True)
# ok = ttk.Button(widgets_frame, text="Okay")

one.grid(column=0, row=12)
two.grid(column=1, row=12)
three.grid(column=2, row=12)
# ok.grid(column=2, row=13)





root.mainloop()