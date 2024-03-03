import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from ttkthemes import ThemedTk
import openpyxl
import pandas as pd

current_sheet = ""  # Global variable to store the selected sheet name
widget_name_mapping = {}  # Global variable to store the mapping of widgets
column_to_widget_mapping = {} 

def load_data(file_path):
    global current_sheet, widget_name_mapping
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet_names = workbook.sheetnames
        
        if not current_sheet or current_sheet not in sheet_names:
            # If current_sheet is empty or not in sheet_names, set it to the first sheet
            current_sheet = sheet_names[0]
            
        # Update the Combobox with sheet names
        sheet_dropdown["values"] = sheet_names
        sheet_dropdown.set(current_sheet)  # Set the selected sheet in the Combobox

        sheet = workbook[current_sheet]
        list_values = list(sheet.iter_rows(values_only=True))
        columns = list_values[0]  # Extract the column headings from the first row of the selected sheet

        # Update the TreeView's headings with the column headings
        update_treeview_headings(columns)

        widget_name_mapping = {}
        column_to_widget_mapping = {}  # Reverse mapping: Column headings to widget names

        for idx, col_name in enumerate(columns):
            if idx < 6:
                widget_name = "widget_" + str(idx)
                widget_name_mapping[widget_name] = col_name
                column_to_widget_mapping[col_name] = widget_name

    except Exception as e:
        print(f"Error loading Excel file: {e}")

def on_sheet_select(event):
    global current_sheet
    selected_sheet = sheet_dropdown.get()
    if selected_sheet != current_sheet:
        current_sheet = selected_sheet
        load_data("db.xlsx")  # Load data for the selected sheet

root = tk.Tk()
root.title("Excel Data Viewer")

# Sheet Combobox
sheet_var = tk.StringVar()
sheet_dropdown = ttk.Combobox(root, textvariable=sheet_var, state="readonly")
sheet_dropdown.grid(row=4, column=5, padx=10, pady=5)
sheet_dropdown.bind("<<ComboboxSelected>>", on_sheet_select)

# Load initial data
load_data("db.xlsx")



def insert_row():
    if not column_to_widget_mapping:
        print("Error: No mapping available.")
    # Retrieve data from the widgets
    r1 = month.get()
    r2 = obr_number.get()
    r3 = category_dropdown.get()
    r4 = sub_category.get()
    r5 = brand.get()
    r6 = price.get()
    r7 = notes.get()

    # Get the selected sheet from the workbook
    path = file_path
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[current_sheet]
    
    # Insert the row to the selected sheet in the workbook
    row_values = [r1, r2, r3, r4, r5, r6, r7]
    sheet.append(row_values)
    workbook.save(path)

    # displaying the inserted row on the UI (treeView)
    treeView.insert('', tk.END, values=row_values)
    
    # Get the ID of the last inserted row in the TreeView
    last_row_id = treeView.get_children()[-1]

    # clear the values after inserting the new row then resetting the values to default
    r1.delete(0, "end")
    r2.delete(0, "end")
    r3.set(status_list[0])
    r4.delete(0, "end")
    r5.delete(0, "end")
    r6.delete(0, "end")
    r7.delete(0, "end")

    # Highlight or select the last inserted row in the TreeView
    treeView.selection_set(last_row_id)
    # Scroll the TreeView to make sure the last inserted row is visible
    treeView.see(last_row_id)
    
    # returns the focus to the check_date widget after inserting the new row
    check_date.focus_set()
    
    # Change the style of the "Add" button when it is selected using Tab
    style.map("Custom.TButton",
              foreground=[("active", "white"), ("!active", "black")],
              background=[("active", "blue"), ("!active", "SystemButtonFace")])


# Create the main window
root = ThemedTk(theme="radiance") #tk.Tk()
root.title('GSO Records App by tEppy™')

# Update labels of widgets based on column headings and get the widget_name_mapping
# sheet_names = load_data(file_path)  # Get the sheet names list

################## Main Frame ##################
'''
    outer_frame.pack() makes the app responsive.
    Since this is the main widget, adjusting the size of the UI/app will keep
    the components centered
'''
outer_frame = ttk.Frame(root) # parent widget
outer_frame.pack()


################## Widgets Frame ##################
''' 
    The user-input form area.
    All other ttk widgets should have "widgets_frame" as their parent
    since they are supposed to be "grouped-together"... except from, of course, the
    widgets_frame for its root will be the outer_frame (see above).
'''

### col0, row0 of the root_frame : Enclosure Widget for all other input widgets
widgets_frame = ttk.LabelFrame(outer_frame, text="Widgets Frame")
widgets_frame.grid(row=0, column=0, padx=20, pady=10) # padding on x & y axis

w00 = ttk.Label(widgets_frame, )
w00.grid(row=0, column=0)

w01 = ttk.Label(widgets_frame, )
w01.grid(row=0, column=1)

w02 = ttk.Label(widgets_frame, )
w02.grid(row=0, column=2)

w03 = ttk.Label(widgets_frame, )
w03.grid(row=0, column=3)

w04 = ttk.Label(widgets_frame, )
w04.grid(row=0, column=4)

w05 = ttk.Label(widgets_frame, )
w05.grid(row=0, column=5)

w06 = ttk.Label(widgets_frame, )
w06.grid(row=0, column=6)


# Read Excel data from a specific sheet (replace 'Sheet1' with your actual sheet name)
df = pd.read_excel('db.xlsx', sheet_name='Lists(DoNotSelectThis)')
# setting variables for dropdown options
asset_type_list = df['AssetType'].tolist()
chairs_list = df['List_Chairs'].tolist()
departments_list = df['Departments'].tolist()
kitchen_list = df['List_Kitchen'].tolist()
others_list = df['List_Others'].tolist()
peripherals_list = df['List_ComputerPeripherals'].tolist()
racks_list = df['List_Racks'].tolist()
tables_list = df['List_Tables'].tolist()


### actual widgets
month = ttk.Entry(widgets_frame, width=7)
# check_date.insert(0, "help text here")
# check_date.bind("<FocusIn>", lambda e: check_date.delete('0', 'end'))
month.grid(row=1, column=0, padx=5, pady=(0,5), sticky="ew")

obr_number = ttk.Entry(widgets_frame,width=13)
obr_number.grid(row=1, column=1, padx=5, pady=(0,5), sticky="ew")

# category_list = asset_type_list
# category_dropdown = ttk.Combobox(widgets_frame, values=category_list, state='readonly')
# category_dropdown.current(0) # default selected value
# category_dropdown.grid(row=1, column=2, padx=5, pady=(0,5), sticky="ew")

def on_category_selected(event):
    selected_category = category_var.get()
    if selected_category == "Computer":
        sub_category_combobox['values'] = peripherals_list
        sub_category_combobox.set(peripherals_list[0])  # Set default value
        sub_category_combobox['state'] = 'readonly'
        sub_category_combobox.grid(row=1, column=3, padx=5, pady=(0, 5), sticky="ew")
    elif selected_category == "Kitchen-related":
        sub_category_combobox['values'] = kitchen_list
        sub_category_combobox.set(kitchen_list[0])  # Set default value
        sub_category_combobox['state'] = 'readonly'
        sub_category_combobox.grid(row=1, column=3, padx=5, pady=(0, 5), sticky="ew")
    elif selected_category == "Table":
        sub_category_combobox['values'] = tables_list
        sub_category_combobox.set(tables_list[0])  # Set default value
        sub_category_combobox['state'] = 'readonly'
        sub_category_combobox.grid(row=1, column=3, padx=5, pady=(0, 5), sticky="ew")
    elif selected_category == "Chair":
        sub_category_combobox['values'] = chairs_list
        sub_category_combobox.set(chairs_list[0])  # Set default value
        sub_category_combobox['state'] = 'readonly'
        sub_category_combobox.grid(row=1, column=3, padx=5, pady=(0, 5), sticky="ew")
    elif selected_category == "Rack":
        sub_category_combobox['values'] = racks_list
        sub_category_combobox.set(racks_list[0])  # Set default value
        sub_category_combobox['state'] = 'readonly'
        sub_category_combobox.grid(row=1, column=3, padx=5, pady=(0, 5), sticky="ew")
    elif selected_category == "Others":
        sub_category_combobox['values'] = others_list
        sub_category_combobox.set(others_list[0])  # Set default value
        sub_category_combobox['state'] = 'readonly'
        sub_category_combobox.grid(row=1, column=3, padx=5, pady=(0, 5), sticky="ew")
    else:
        sub_category_combobox['state'] = 'disabled'
        sub_category_combobox.grid(row=1, column=3, padx=5, pady=(0, 5), sticky="ew")

# Category Combobox
category_var = tk.StringVar()
category_combobox = ttk.Combobox(widgets_frame, textvariable=category_var, values=asset_type_list, state='readonly')
category_combobox.set("Select Category")
category_combobox.grid(row=1, column=2, padx=5, pady=(0, 5), sticky="ew")
category_combobox.bind("<<ComboboxSelected>>", on_category_selected)

# Subcategory Combobox (initially hidden)
sub_category_combobox = ttk.Combobox(widgets_frame)
sub_category_combobox.grid(row=1, column=3, padx=5, pady=(0, 5), sticky="ew")








brand = ttk.Entry(widgets_frame, width=20)
brand.grid(row=1, column=4,padx=5, pady=(0,5), sticky="ew")

price = ttk.Entry(widgets_frame, width=15)
price.grid(row=1, column=5,padx=5, pady=(0,5), sticky="ew")

notes = ttk.Entry(widgets_frame, width=25)
notes.grid(row=1, column=6, padx=5, pady=(0,5), sticky="ew")

# Select File Button
# btn_load = ttk.Button(widgets_frame, text="Load Excel File", command=load_excel_file, takefocus=0)
# btn_load.grid(row=0, column=7, padx=5, pady=(0, 5), sticky="ew")

# Select Sheet Button (Inside Widgets Frame)
def select_sheet():
    selected_item = treeView.focus()
    if selected_item:
        item_values = treeView.item(selected_item, "values")
        selected_sheet_name = item_values[0]
        print(f"Selected sheet: {selected_sheet_name}")

# Dropdown menu for selecting sheets
sheet_dropdown = ttk.Combobox(widgets_frame, state="readonly")
sheet_dropdown.grid(row=4, column=5, padx=10, pady=5)
sheet_dropdown.bind("<<ComboboxSelected>>", on_sheet_select)

# Insert Row button (Inside Widgets Frame)
btn_row = ttk.Button(widgets_frame, text="Insert", command=lambda: insert_row(),
                    takefocus=1)
btn_row.grid(row=4, column=3, sticky="nsew")


################## TreeView / Excel LabelFrame ####################################
### This is where the preview of the excel file's data will be displayed

### Display Frame
treeFrame = ttk.Frame(outer_frame, takefocus=0)
treeFrame.grid(row=5, column=0, pady=10)
treeScroll = ttk.Scrollbar(treeFrame)
treeScroll.pack(side="right", fill="y") # this sets the scrollbar to the right side of the frame,
                                        # covering its whole height
cols = ("Month", "OBR #", "Category","Sub-category", "Brand", "Price", "Notes") # column of the preview related to the excel file
treeView = ttk.Treeview(treeFrame, show="headings", 
                        yscrollcommand=treeScroll.set, columns=cols, height=15)
# these set the width of the columns specifically
# COLUMN NAMES SHOULD MATCH THOSE INDICATED ON THE EXCEL FILE
treeView.column("Month", width=100)
treeView.column("OBR #", width=100)
treeView.column("Category", width=100)
treeView.column("Sub-category", width=200)
treeView.column("Brand", width=100)
treeView.column("Price", width=100)
treeView.column("Notes", width=80)
treeView.pack()
treeScroll.config(command=treeView.yview) # this line attaches the treeScroll widget to the treeView, scrolling vertically


# Event Listener function highlighting selected items on the treeView list
def selected():
    print(listbox.get(listbox.curselection()[0]))
    
treeView.bind("<<ListboxSelect>>", lambda x: selected())

# ### switch (dark/light)
# def toggle_mode():
#     if mode_switch.instate(["selected"]):
#         style.theme_use("forest-light")
#     else:
#         style.theme_use("forest-dark")

# mode_switch = ttk.Checkbutton(outer_frame, text="Mode", style="Switch",
#     command=toggle_mode, takefocus=0) # this triggers the toggle_mode function above
# mode_switch.grid(row=6, column=0, padx=5, pady=10, sticky="nsew")

month.focus_set()

root.mainloop()
